import pytest
import tempfile
from myapp.models import Issue, Analysis
from django.urls import reverse
from django.contrib.auth.models import User
from myapp.models import Issue
from django.utils.timezone import now
from django.test import Client
from django.core.files.uploadedfile import SimpleUploadedFile



@pytest.mark.django_db
@pytest.mark.parametrize(
    "issues_data, expected_status_code, expected_titles_in_response, expected_statuses_in_response",
    [
        #  issue is present
        (
            [{"title": "Graffiti on Wall", "status": "open"}],
            200,
            ["Graffiti on Wall"],
            ["open"]
        ),
        # No issues present (should be empty page or no issue listed)
        (
            [],
            200,
            [],
            []
        ),
    ]
)
def test_view_issues(client, issues_data, expected_status_code, expected_titles_in_response, expected_statuses_in_response):
    # Create issues based on input data
    for issue in issues_data:
        Issue.objects.create(title=issue["title"], status=issue["status"])

    # Perform the GET request to the view
    response = client.get(reverse("view_issues"))

    # Assert the response status code
    assert response.status_code == expected_status_code

    # Check that each expected title appears in the response content
    for title in expected_titles_in_response:
        assert title in response.content.decode()

    # Check that each expected status appears in the response content
    for status in expected_statuses_in_response:
        assert status in response.content.decode()


@pytest.mark.django_db
def test_home_view(client):
    response = client.get(reverse("home"))
    assert response.status_code == 200
    assert "issues" in response.context


@pytest.mark.django_db
def test_issue_map_view(client):
    response = client.get(reverse("issue_map"))
    assert response.status_code == 200
    assert "issues" in response.context


@pytest.mark.django_db
def test_report_issue_authenticated(client, django_user_model):
    user = django_user_model.objects.create_user(
        username="testuser", password="password"
    )
    client.login(username="testuser", password="password")
    response = client.get(reverse("report_issue"))
    assert response.status_code == 200
    assert "form" in response.context


@pytest.mark.django_db
def test_report_issue_unauthenticated(client):
    response = client.get(reverse("report_issue"))
    assert response.status_code == 302  # Redirects to login


@pytest.mark.django_db
def test_submitted_issue_view(client):
    response = client.get(reverse("issue_submitted"))
    assert response.status_code == 200


@pytest.mark.django_db
def test_assigned_issues_staff(client, django_user_model):
    staff_user = django_user_model.objects.create_user(
        username="staffuser", password="password", is_staff=True
    )
    client.login(username="staffuser", password="password")
    response = client.get(reverse("assigned_issues"))
    assert response.status_code == 200


@pytest.mark.django_db
def test_assigned_issues_non_staff(client, django_user_model):
    user = django_user_model.objects.create_user(
        username="testuser", password="password"
    )
    client.login(username="testuser", password="password")
    response = client.get(reverse("assigned_issues"))
    assert response.status_code == 403  # Forbidden


@pytest.mark.django_db
@pytest.mark.parametrize("username, password, expected_status_code, expected_redirect", [
    ('testuser', 'testpassword', 302, 'home'),  # Valid credentials, should redirect to home
    ('testuser', 'wrongpassword', 200, None),  # Invalid password, should stay on login page
    ('wronguser', 'testpassword', 200, None),  # Invalid username, should stay on login page

])
def test_login_user(client, username, password, expected_status_code, expected_redirect):

    # Create a test user
    User.objects.create_user(username='testuser', password='testpassword')

    # Post login data
    login_url = reverse('login')  # Adjust the URL name if necessary
    response = client.post(login_url, {'username': username, 'password': password})

    # Check status code and redirection
    assert response.status_code == expected_status_code

    if expected_redirect:
        # Check that we redirect to the homepage after successful login
        assert response.url.endswith(reverse(expected_redirect))
    else:
        # If no redirection, check the login page contains the login form
        assert 'name="username"' in response.content.decode()
        assert 'name="password"' in response.content.decode()


@pytest.mark.django_db
@pytest.mark.parametrize("username, password1, password2, expected_status_code, expected_redirect, expected_error_message", [
    ('newuser', 'password', 'password', 302, 'home', None),  # Valid registration, should redirect to home
    ('newuser', 'password', 'differentpassword', 200, None, 'password2'),  # Password mismatch, should stay on the page
    ('existinguser', 'password', 'password', 200, None, 'username'),  # Username already exists, should stay on the page
])
def test_register_user(client, username, password1, password2, expected_status_code, expected_redirect, expected_error_message):
    # If username already exists, create the user first
    if expected_error_message == 'username':
        User.objects.create_user(username='existinguser', password='password')

    # Post registration data
    response = client.post(
        reverse("register"),
        {"username": username, "password1": password1, "password2": password2},
    )

    # Check the status code
    assert response.status_code == expected_status_code


@pytest.mark.django_db
def test_logout_view(client, django_user_model):
    user = django_user_model.objects.create_user(
        username="testuser", password="password"
    )
    client.login(username="testuser", password="password")
    response = client.get(reverse("logout"))
    assert response.status_code == 302  # Redirect to home


@pytest.mark.django_db
def test_issue_management_staff(client, django_user_model):
    staff_user = django_user_model.objects.create_user(
        username="staffuser", password="password", is_staff=True
    )
    client.login(username="staffuser", password="password")
    response = client.get(reverse("issue_management"))
    assert response.status_code == 200


@pytest.mark.django_db
def test_issue_management_non_staff(client, django_user_model):
    user = django_user_model.objects.create_user(
        username="testuser", password="password"
    )
    client.login(username="testuser", password="password")
    response = client.get(reverse("issue_management"))
    assert response.status_code == 403  # Forbidden


@pytest.mark.django_db
def test_update_issue_as_staff(client):
    # Create a staff user and an issue to update
    user = User.objects.create_user(username="testuser", password="password", is_staff=True)
    issue = Issue.objects.create(
        title="Old Title", 
        description="Old description", 
        status="open", 
        assigned_to=user, 
        location="Old Location", 
        contact_email="old_email@example.com"
    )
    
    # Log in as the staff user
    client.login(username="testuser", password="password")
    
    # URL to update the issue
    url = reverse("update_issue", args=[issue.id])
    
    # Data to update the issue with
    updated_data = {
        "title": "New Title",
        "description": "New description",
        "status": "closed",
        "assigned_to": user.id,  # Ensure it's a valid user ID
        "location": "New Location",
        "contact_email": "new_email@example.com",
    }

    # Send POST request to update the issue
    response = client.post(url, updated_data, follow=True)

    # Assert the response is a redirect or success
    assert response.status_code == 200  # Ensure form submission is successful
    
    # Refresh the issue and assert it was updated
    issue.refresh_from_db()
    assert issue.title == "New Title"
    assert issue.description == "New description"
    assert issue.status == "closed"
    assert issue.location == "New Location"
    assert issue.contact_email == "new_email@example.com"

@pytest.mark.django_db
def test_delete_issue_staff(client, django_user_model):
    staff_user = django_user_model.objects.create_user(
        username="staffuser", password="password", is_staff=True
    )
    # Create Issue
    issue = Issue.objects.create(title="Test Issue", status="open", created_at=now())
    client.login(username="staffuser", password="password")
    # Delete Issue
    response = client.post(reverse("delete_issue", args=[issue.id]))

    # Assert Issue created, no longer exists
    assert response.status_code == 302  # Redirect
    assert not Issue.objects.filter(id=issue.id).exists()


import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.contrib.auth.models import User
from myapp.models import Issue

@pytest.mark.django_db
def test_report_issue():
    client = Client()
    user = User.objects.create_user(username="testuser", password="password")
    url = reverse("report_issue")

    client.login(username="testuser", password="password")

    # test image with EXIF data available
    image_path = 'myapp\\tests\\media\\graffiti1.jpg'  

    with open(image_path, 'rb') as img:
        uploaded_image = SimpleUploadedFile(
            "graffiti1.jpg", img.read(), content_type="image/jpeg"
        )

        data = {
            "title": "Pothole on Main Street",
            "description": "Large pothole causing issues",
            "image": uploaded_image,
        }

        response = client.post(url, data, follow=True)

        # Assert that an issue is created
        assert Issue.objects.count() == 1  # Issue should be created

        # Query the created issue
        issue = Issue.objects.first()

        # Assert that key fields are correct
        assert issue.title == "Pothole on Main Street"
        assert issue.description == "Large pothole causing issues"

        # Assert GPS has been extracted
        assert issue.latitude == 51.18596260343217
        assert issue.longitude == 0.279503587196468

        # Other assertions
        assert response.status_code == 200
        assert response.redirect_chain[0][0] == reverse("issue_submitted")  # Redirect expected


import pytest
from django.urls import reverse
from django.contrib.auth.models import User
from myapp.models import Analysis, Issue
from openpyxl import load_workbook
from io import BytesIO

@pytest.mark.django_db
def test_export_analysis_report(client):
    # Create a user and an issue
    user = User.objects.create_user(username="staffuser", password="password", is_staff=True)
    issue = Issue.objects.create(title="Test Issue", description="Test issue", status="open")
    
    # Create an analysis record
    analysis = Analysis.objects.create(
        issue=issue,
        staff_member=user,
        comment="Test comment",
        estimated_cost=50.0,
        estimated_time=3,
        energy_required="High",
        priority_level=8,
        recommended_action="Repair",
        environmental_impact=75
    )
    
    # Log in as the user
    client.login(username="staffuser", password="password")
    
    # Call the export view
    url = reverse("export_analysis_report")
    response = client.get(url)
    
    # Check if the response is an Excel file
    assert response.status_code == 200
    assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    # Load the response content into a workbook
    file_content = BytesIO(response.content)
    wb = load_workbook(file_content)
    ws = wb.active
    
    # Check the first row (headers)
    headers = ['Issue ID', 'Issue Title', 'Analysis Comment', 'Left By', 'Estimated Cost (£)',
               'Estimated Time (Days)', 'Energy Required', 'Priority Level (1-10)', 'Recommended Action', 'Environmental Impact (1-100)']
    
    for idx, header in enumerate(headers, 1):
        assert ws.cell(row=1, column=idx).value == header
    
    # Check the second row (data)
    assert ws.cell(row=2, column=1).value == issue.id
    assert ws.cell(row=2, column=2).value == issue.title
    assert ws.cell(row=2, column=3).value == analysis.comment
    assert ws.cell(row=2, column=4).value == user.username
    assert ws.cell(row=2, column=5).value == analysis.estimated_cost
    assert ws.cell(row=2, column=6).value == analysis.estimated_time
    assert ws.cell(row=2, column=7).value == analysis.energy_required
    assert ws.cell(row=2, column=8).value == analysis.priority_level
    assert ws.cell(row=2, column=9).value == analysis.recommended_action
    assert ws.cell(row=2, column=10).value == analysis.environmental_impact


